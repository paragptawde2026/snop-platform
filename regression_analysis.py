import warnings
warnings.filterwarnings('ignore')

import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet
from sklearn.preprocessing import PolynomialFeatures, StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.svm import SVR
from sklearn.model_selection import cross_val_score
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("Reading Excel file...")
df = pd.read_excel(r'D:\Furnace PIMS\Regression Database\PK1\coilsim_reg_H1-H6 2.xlsx')

# X variables: columns B-H (indices 1-7), skipping Unnamed:0 which is a date
x_cols = list(df.columns[1:8])
# Y variables: columns I-Q (indices 8-16), excluding constants (std=0)
all_y_cols = list(df.columns[8:17])
y_cols = [c for c in all_y_cols if df[c].std() > 0]

print(f"X variables ({len(x_cols)}): {x_cols}")
print(f"Y variables ({len(y_cols)}): {y_cols}")
print(f"Skipped constant Y columns: {[c for c in all_y_cols if c not in y_cols]}")

# Drop rows with any NaN in X or Y
data = df[x_cols + y_cols].dropna()
print(f"\nData shape after dropping NaN: {data.shape}")

X = data[x_cols].values
Y = data[y_cols]

def compute_metrics(y_true, y_pred, n_features):
    r2 = r2_score(y_true, y_pred)
    n = len(y_true)
    adj_r2 = 1 - (1 - r2) * (n - 1) / (n - n_features - 1)
    rmse = np.sqrt(mean_squared_error(y_true, y_pred))
    mae = mean_absolute_error(y_true, y_pred)
    mask = y_true != 0
    mape = np.mean(np.abs((y_true[mask] - y_pred[mask]) / y_true[mask])) * 100 if mask.any() else np.nan
    return {'R2': r2, 'Adj_R2': adj_r2, 'RMSE': rmse, 'MAE': mae, 'MAPE': mape}

scaler = StandardScaler()
X_scaled = scaler.fit_transform(X)
n_features = X.shape[1]

models = {
    'Linear Regression': LinearRegression(),
    'Ridge': Ridge(alpha=1.0),
    'Lasso': Lasso(alpha=0.001, max_iter=10000),
    'Elastic Net': ElasticNet(alpha=0.001, l1_ratio=0.5, max_iter=10000),
    'Polynomial (deg2)': Pipeline([('poly', PolynomialFeatures(degree=2, include_bias=False)), ('lr', LinearRegression())]),
    'Random Forest': RandomForestRegressor(n_estimators=100, random_state=42, n_jobs=-1),
    'Gradient Boosting': GradientBoostingRegressor(n_estimators=100, random_state=42),
    'SVR': SVR(kernel='rbf', C=10, epsilon=0.1),
}

# Scale-sensitive models use scaled X
scale_sensitive = {'Lasso', 'Ridge', 'Elastic Net', 'SVR', 'Polynomial (deg2)'}

all_results = {}
best_models_info = {}
actual_vs_predicted = {}

for y_col in y_cols:
    print(f"\nProcessing: {y_col}")
    y = data[y_col].values
    col_results = []

    for model_name, model in models.items():
        X_input = X_scaled if model_name in scale_sensitive else X
        try:
            model.fit(X_input, y)
            y_pred = model.predict(X_input)
            metrics = compute_metrics(y, y_pred, n_features)

            # Cross-validation
            try:
                cv_scores = cross_val_score(model, X_input, y, cv=5, scoring='r2', n_jobs=-1)
                metrics['CV_R2'] = cv_scores.mean()
                metrics['CV_R2_std'] = cv_scores.std()
            except Exception:
                metrics['CV_R2'] = np.nan
                metrics['CV_R2_std'] = np.nan

            metrics['model_name'] = model_name
            metrics['y_col'] = y_col
            col_results.append(metrics)
            print(f"  {model_name}: R2={metrics['R2']:.4f}, RMSE={metrics['RMSE']:.4f}")

        except Exception as e:
            print(f"  {model_name}: FAILED - {e}")
            col_results.append({
                'model_name': model_name, 'y_col': y_col,
                'R2': np.nan, 'Adj_R2': np.nan, 'RMSE': np.nan,
                'MAE': np.nan, 'MAPE': np.nan, 'CV_R2': np.nan, 'CV_R2_std': np.nan
            })

    all_results[y_col] = col_results

    # Find best model
    valid = [r for r in col_results if not np.isnan(r['R2'])]
    if valid:
        best = max(valid, key=lambda r: r['R2'])
        best_model_name = best['model_name']
        best_models_info[y_col] = best.copy()

        # Get feature importances / coefficients for best model
        model_obj = models[best_model_name]
        X_input = X_scaled if best_model_name in scale_sensitive else X
        model_obj.fit(X_input, y)
        y_pred_best = model_obj.predict(X_input)
        best_models_info[y_col]['predictions'] = y_pred_best

        if hasattr(model_obj, 'feature_importances_'):
            importances = model_obj.feature_importances_
        elif hasattr(model_obj, 'coef_'):
            importances = model_obj.coef_
        elif hasattr(model_obj, 'named_steps'):
            step = list(model_obj.named_steps.values())[-1]
            if hasattr(step, 'coef_'):
                importances = step.coef_
            else:
                importances = [np.nan] * n_features
        else:
            importances = [np.nan] * n_features

        # For poly, we have more features, store only original feature importances summary
        if len(importances) != len(x_cols):
            best_models_info[y_col]['importances'] = dict(zip(x_cols, [np.nan] * len(x_cols)))
        else:
            best_models_info[y_col]['importances'] = dict(zip(x_cols, importances))

        actual_vs_predicted[y_col] = {'actual': y, 'predicted': y_pred_best}

print("\nAll models built successfully!")

# ============================================================
# CREATE EXCEL OUTPUT
# ============================================================
print("\nCreating Excel output file...")

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Color definitions
HDR_FILL = PatternFill('solid', start_color='BDD7EE', end_color='BDD7EE')
BEST_FILL = PatternFill('solid', start_color='E2EFDA', end_color='E2EFDA')
WORST_FILL = PatternFill('solid', start_color='FCE4D6', end_color='FCE4D6')
WHITE_FILL = PatternFill('solid', start_color='FFFFFF', end_color='FFFFFF')
ALT_FILL = PatternFill('solid', start_color='F5F5F5', end_color='F5F5F5')

HDR_FONT = Font(name='Arial', bold=True, size=10)
BODY_FONT = Font(name='Arial', size=10)
TITLE_FONT = Font(name='Arial', bold=True, size=12)

thin = Side(style='thin', color='CCCCCC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, fill=None):
    cell.font = HDR_FONT
    cell.fill = fill or HDR_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER

def style_body(cell, fill=None):
    cell.font = BODY_FONT
    cell.fill = fill or WHITE_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER

def auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ''
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

def format_r2(val):
    return round(float(val), 4) if val is not None and not (isinstance(val, float) and np.isnan(val)) else 'N/A'

def format_metric(val):
    return round(float(val), 4) if val is not None and not (isinstance(val, float) and np.isnan(val)) else 'N/A'

# ---- SHEET 1: Summary ----
ws_sum = wb.create_sheet('Summary')
ws_sum['A1'] = 'Regression Analysis Summary - All Models vs All Y Variables'
ws_sum['A1'].font = TITLE_FONT
ws_sum.row_dimensions[1].height = 20

# Headers row 3
headers = ['Y Variable', 'Best Model', 'Best R²', 'Best RMSE', 'Best MAE', 'Best CV R²'] + \
          [f'{m}\nR²' for m in models.keys()] + [f'{m}\nRMSE' for m in models.keys()]
for col_idx, h in enumerate(headers, 1):
    cell = ws_sum.cell(row=3, column=col_idx, value=h)
    style_header(cell)

model_names = list(models.keys())
for row_idx, y_col in enumerate(y_cols, 4):
    results_map = {r['model_name']: r for r in all_results[y_col]}
    valid = {k: v for k, v in results_map.items() if not np.isnan(v.get('R2', np.nan))}

    if valid:
        best_name = max(valid, key=lambda k: valid[k]['R2'])
        worst_name = min(valid, key=lambda k: valid[k]['R2'])
        best_r = valid[best_name]
    else:
        best_name = worst_name = ''
        best_r = {}

    row_data = [
        y_col,
        best_name,
        format_r2(best_r.get('R2')),
        format_metric(best_r.get('RMSE')),
        format_metric(best_r.get('MAE')),
        format_metric(best_r.get('CV_R2')),
    ]
    for m in model_names:
        row_data.append(format_r2(results_map.get(m, {}).get('R2')))
    for m in model_names:
        row_data.append(format_metric(results_map.get(m, {}).get('RMSE')))

    fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_sum.cell(row=row_idx, column=col_idx, value=val)
        style_body(cell, fill=fill)

ws_sum.freeze_panes = 'A4'
auto_fit_columns(ws_sum)

# ---- SHEET 2: Model_Comparison ----
ws_mc = wb.create_sheet('Model_Comparison')
ws_mc['A1'] = 'Detailed Metrics - All Models for Each Y Variable'
ws_mc['A1'].font = TITLE_FONT

mc_headers = ['Y Variable', 'Model', 'R²', 'Adj R²', 'RMSE', 'MAE', 'MAPE (%)', 'CV R²', 'CV R² Std', 'Rank']
for col_idx, h in enumerate(mc_headers, 1):
    cell = ws_mc.cell(row=3, column=col_idx, value=h)
    style_header(cell)

mc_row = 4
for y_col in y_cols:
    results = all_results[y_col]
    valid = [r for r in results if not np.isnan(r.get('R2', np.nan))]
    if valid:
        sorted_valid = sorted(valid, key=lambda r: r['R2'], reverse=True)
        rank_map = {r['model_name']: i + 1 for i, r in enumerate(sorted_valid)}
        best_name = sorted_valid[0]['model_name'] if sorted_valid else ''
        worst_name = sorted_valid[-1]['model_name'] if len(sorted_valid) > 1 else ''
    else:
        rank_map = {}
        best_name = worst_name = ''

    for r in results:
        rank = rank_map.get(r['model_name'], len(results))
        row_data = [
            y_col, r['model_name'],
            format_r2(r.get('R2')), format_r2(r.get('Adj_R2')),
            format_metric(r.get('RMSE')), format_metric(r.get('MAE')),
            format_metric(r.get('MAPE')), format_metric(r.get('CV_R2')),
            format_metric(r.get('CV_R2_std')), rank
        ]
        if r['model_name'] == best_name:
            fill = BEST_FILL
        elif r['model_name'] == worst_name:
            fill = WORST_FILL
        else:
            fill = WHITE_FILL

        for col_idx, val in enumerate(row_data, 1):
            cell = ws_mc.cell(row=mc_row, column=col_idx, value=val)
            style_body(cell, fill=fill)
        mc_row += 1

ws_mc.freeze_panes = 'A4'
auto_fit_columns(ws_mc)

# ---- SHEET 3: Best_Models ----
ws_bm = wb.create_sheet('Best_Models')
ws_bm['A1'] = 'Best Model Per Y Variable with Feature Importances/Coefficients'
ws_bm['A1'].font = TITLE_FONT

bm_headers = ['Y Variable', 'Best Model', 'R²', 'Adj R²', 'RMSE', 'MAE', 'MAPE (%)', 'CV R²'] + x_cols
for col_idx, h in enumerate(bm_headers, 1):
    cell = ws_bm.cell(row=3, column=col_idx, value=h)
    style_header(cell)

for row_idx, y_col in enumerate(y_cols, 4):
    info = best_models_info.get(y_col, {})
    importances = info.get('importances', {})
    row_data = [
        y_col, info.get('model_name', ''),
        format_r2(info.get('R2')), format_r2(info.get('Adj_R2')),
        format_metric(info.get('RMSE')), format_metric(info.get('MAE')),
        format_metric(info.get('MAPE')), format_metric(info.get('CV_R2')),
    ] + [format_metric(importances.get(x)) for x in x_cols]

    fill = BEST_FILL if row_idx % 2 == 0 else PatternFill('solid', start_color='D8F0D0', end_color='D8F0D0')
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_bm.cell(row=row_idx, column=col_idx, value=val)
        style_body(cell, fill=BEST_FILL)

ws_bm.freeze_panes = 'A4'
auto_fit_columns(ws_bm)

# ---- SHEET 4: Actual_vs_Predicted ----
ws_avp = wb.create_sheet('Actual_vs_Predicted')
ws_avp['A1'] = 'Actual vs Predicted Values - Best Model Per Y Variable'
ws_avp['A1'].font = TITLE_FONT

col_offset = 1
avp_col_map = {}
for y_col in y_cols:
    avp_col_map[y_col] = col_offset
    ws_avp.cell(row=3, column=col_offset, value=f'{y_col}\nActual').font = HDR_FONT
    ws_avp.cell(row=3, column=col_offset + 1, value=f'{y_col}\nPredicted').font = HDR_FONT
    ws_avp.cell(row=3, column=col_offset + 2, value=f'{y_col}\nResidual').font = HDR_FONT
    for c in range(col_offset, col_offset + 3):
        cell = ws_avp.cell(row=3, column=c)
        style_header(cell)
    col_offset += 4

avp_data = actual_vs_predicted
n_rows = len(data)
for row_idx in range(n_rows):
    col_offset = 1
    fill = WHITE_FILL if row_idx % 2 == 0 else ALT_FILL
    for y_col in y_cols:
        avp = avp_data.get(y_col, {})
        actual_arr = avp.get('actual', [])
        pred_arr = avp.get('predicted', [])
        if row_idx < len(actual_arr):
            act_val = round(float(actual_arr[row_idx]), 6)
            pred_val = round(float(pred_arr[row_idx]), 6)
            resid_val = round(act_val - pred_val, 6)
        else:
            act_val = pred_val = resid_val = ''

        for c, val in enumerate([act_val, pred_val, resid_val], col_offset):
            cell = ws_avp.cell(row=row_idx + 4, column=c, value=val)
            style_body(cell, fill=fill)
        col_offset += 4

ws_avp.freeze_panes = 'A4'
auto_fit_columns(ws_avp)

# ---- ONE SHEET PER Y VARIABLE ----
for y_col in y_cols:
    safe_name = y_col[:31].replace('/', '_').replace('\\', '_').replace('*', '').replace('?', '').replace('[', '').replace(']', '').replace(':', '')
    ws_y = wb.create_sheet(safe_name)
    ws_y[f'A1'] = f'Model Comparison for: {y_col}'
    ws_y['A1'].font = TITLE_FONT

    y_headers = ['Model', 'R²', 'Adj R²', 'RMSE', 'MAE', 'MAPE (%)', 'CV R²', 'CV R² Std', 'Rank']
    for col_idx, h in enumerate(y_headers, 1):
        cell = ws_y.cell(row=3, column=col_idx, value=h)
        style_header(cell)

    results = all_results[y_col]
    valid = [r for r in results if not np.isnan(r.get('R2', np.nan))]
    sorted_valid = sorted(valid, key=lambda r: r['R2'], reverse=True)
    rank_map = {r['model_name']: i + 1 for i, r in enumerate(sorted_valid)}
    best_name = sorted_valid[0]['model_name'] if sorted_valid else ''
    worst_name = sorted_valid[-1]['model_name'] if len(sorted_valid) > 1 else ''

    sorted_results = sorted(results, key=lambda r: (rank_map.get(r['model_name'], 99)))
    for row_idx, r in enumerate(sorted_results, 4):
        rank = rank_map.get(r['model_name'], len(results))
        row_data = [
            r['model_name'],
            format_r2(r.get('R2')), format_r2(r.get('Adj_R2')),
            format_metric(r.get('RMSE')), format_metric(r.get('MAE')),
            format_metric(r.get('MAPE')), format_metric(r.get('CV_R2')),
            format_metric(r.get('CV_R2_std')), rank
        ]
        if r['model_name'] == best_name:
            fill = BEST_FILL
        elif r['model_name'] == worst_name:
            fill = WORST_FILL
        else:
            fill = WHITE_FILL if row_idx % 2 == 0 else ALT_FILL

        for col_idx, val in enumerate(row_data, 1):
            cell = ws_y.cell(row=row_idx, column=col_idx, value=val)
            style_body(cell, fill=fill)

    ws_y.freeze_panes = 'A4'
    auto_fit_columns(ws_y)

output_path = r'D:\Furnace PIMS\Regression Database\PK1\regression_results.xlsx'
wb.save(output_path)
print(f"\nOutput saved to: {output_path}")
print("\n=== SUMMARY ===")
print(f"X variables: {x_cols}")
print(f"Y variables modeled: {y_cols}")
print(f"Skipped (constant): {[c for c in all_y_cols if c not in y_cols]}")
print(f"Samples used: {len(data)}")
for y_col in y_cols:
    info = best_models_info.get(y_col, {})
    print(f"  {y_col}: Best={info.get('model_name','?')} R2={format_r2(info.get('R2'))} RMSE={format_metric(info.get('RMSE'))}")
